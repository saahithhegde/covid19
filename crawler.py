import requests
import bs4
import openpyxl

#crawler initializer
from openpyxl import Workbook

url="https://www.mohfw.gov.in/"
response=requests.get(url)

soup =bs4.BeautifulSoup(response.text,"html.parser")

tabletags=soup.select('table')

mytable=tabletags[9]

tablerows=mytable.find_all('tr')

#excel initialiser
workbook=Workbook()
sheet1=workbook.active
filename="data.xlsx"

sheet1.cell(row=1,column=1).value="STATES"
sheet1.cell(row=1,column=2).value="ACTIVE"
sheet1.cell(row=1,column=3).value="CURED"
sheet1.cell(row=1,column=4).value="DEATHS"




k=2
active=0
death=0
cured=0

for i in range(1,len(tablerows)-1):
    sheet1.cell(row=k,column=1).value=tablerows[i].find_all('td')[1].get_text()
    sheet1.cell(row=k, column=2).value = tablerows[i].find_all('td')[2].get_text()
    active+=int(tablerows[i].find_all('td')[2].get_text())
    sheet1.cell(row=k, column=3).value = tablerows[i].find_all('td')[3].get_text()
    cured += int(tablerows[i].find_all('td')[3].get_text())
    sheet1.cell(row=k, column=4).value = tablerows[i].find_all('td')[4].get_text()
    death += int(tablerows[i].find_all('td')[4].get_text())
    k+=1

sheet1.cell(row=k,column=1).value="TOTAL"
sheet1.cell(row=k,column=2).value=str(active)
sheet1.cell(row=k,column=3).value=str(cured)
sheet1.cell(row=k,column=4).value=str(death)
workbook.save(filename=filename)